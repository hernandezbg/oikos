from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import LoginView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import TemplateView, CreateView, ListView, UpdateView
from django.urls import reverse_lazy
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Q
from django.utils import timezone
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from decimal import Decimal

from core.models import Movimiento, SaldoMensual, CategoriaIngreso, CategoriaEgreso, Iglesia
from core.forms import MovimientoForm, FiltroMovimientosForm, RegistroForm, CategoriaIngresoForm, CategoriaEgresoForm
from core.forms_google import RegistroIglesiaGoogleForm
from core.utils import formato_pesos, calcular_saldo_mes, generar_reporte_pdf, get_dashboard_data, formato_mes
from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required


class AccesoMovimientosRequiredMixin:
    """
    Mixin que verifica que el usuario tenga acceso a movimientos generales.
    Los usuarios de solo caja no pueden acceder a estas vistas.
    """
    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            if not request.user.is_staff and not request.user.is_superuser:
                if not request.user.iglesia:
                    return redirect('seleccionar_tipo_registro')
                # Si es usuario solo de caja, no tiene acceso
                if not request.user.tiene_acceso_movimientos:
                    messages.warning(request, 'No tienes permisos para acceder a esta sección.')
                    return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)


def home_view(request):
    """
    Vista de landing page - muestra home.html
    """
    return render(request, 'core/home.html')


class CustomLoginView(LoginView):
    template_name = 'core/login.html'

    def dispatch(self, request, *args, **kwargs):
        # Si el usuario ya está autenticado, redirigir al dashboard
        if request.user.is_authenticated:
            return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['app_name'] = 'OIKOS'
        return context


def registro_view(request):
    """
    Vista para registro de nueva iglesia y usuario tesorero
    """
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        form = RegistroForm(request.POST)
        if form.is_valid():
            user = form.save()
            # Iniciar sesión automáticamente
            login(request, user)
            messages.success(
                request,
                f'¡Bienvenido a OIKOS! La iglesia {user.iglesia.nombre} ha sido registrada exitosamente.'
            )
            return redirect('dashboard')
    else:
        form = RegistroForm()

    return render(request, 'core/registro.html', {'form': form})


@login_required
def seleccionar_tipo_registro_view(request):
    """
    Vista para seleccionar si crear iglesia o unirse con código
    """
    # Si ya tiene iglesia, redirigir al dashboard
    if hasattr(request.user, 'iglesia') and request.user.iglesia:
        return redirect('dashboard')

    # Si es admin/staff, puede acceder al dashboard sin iglesia
    if request.user.is_staff or request.user.is_superuser:
        return redirect('dashboard')

    from core.forms_invitacion import SeleccionTipoRegistroForm

    if request.method == 'POST':
        form = SeleccionTipoRegistroForm(request.POST)
        if form.is_valid():
            tipo = form.cleaned_data['tipo']
            if tipo == 'crear_iglesia':
                return redirect('registro_iglesia_google')
            else:
                return redirect('registro_con_codigo')
    else:
        form = SeleccionTipoRegistroForm()

    return render(request, 'core/seleccionar_tipo_registro.html', {
        'form': form,
        'user': request.user
    })


def registro_iglesia_google_view(request):
    """
    Vista para crear nueva iglesia (primer usuario = ADMIN)
    """
    # Si ya tiene iglesia, redirigir al dashboard
    if hasattr(request.user, 'iglesia') and request.user.iglesia:
        return redirect('dashboard')

    if request.user.is_staff or request.user.is_superuser:
        return redirect('dashboard')

    if request.method == 'POST':
        form = RegistroIglesiaGoogleForm(request.POST)
        if form.is_valid():
            # Crear la iglesia
            iglesia = form.save(commit=False)
            iglesia.activa = True
            iglesia.save()

            # Asignar iglesia al usuario y hacerlo ADMIN
            request.user.iglesia = iglesia
            request.user.rol = 'ADMIN'  # El fundador es ADMIN
            request.user.puede_aprobar = True
            request.user.save()

            messages.success(
                request,
                f'¡Bienvenido a OIKOS! La iglesia {iglesia.nombre} ha sido registrada exitosamente. '
                f'Eres el administrador y puedes invitar a otros usuarios.'
            )
            return redirect('dashboard')
    else:
        form = RegistroIglesiaGoogleForm()

    return render(request, 'core/registro_iglesia_google.html', {
        'form': form,
        'user': request.user
    })


@login_required
def registro_con_codigo_view(request):
    """
    Vista para unirse a una iglesia O a una caja chica usando un código de invitación
    """
    from core.forms_invitacion import ValidarCodigoInvitacionForm
    from core.models import CodigoInvitacion, UsuarioCajaChica

    if request.method == 'POST':
        form = ValidarCodigoInvitacionForm(request.POST)
        if form.is_valid():
            codigo_obj = form.codigo_obj

            # CASO 1: Código para iglesia (comportamiento original)
            if codigo_obj.caja_chica is None:
                # Si ya tiene iglesia, no puede usar código de iglesia
                if hasattr(request.user, 'iglesia') and request.user.iglesia:
                    messages.error(request, 'Ya perteneces a una iglesia')
                    return redirect('dashboard')

                # Asignar iglesia y rol al usuario
                request.user.iglesia = codigo_obj.iglesia
                request.user.rol = codigo_obj.rol

                # Permisos según rol
                if codigo_obj.rol in ['ADMIN', 'TESORERO']:
                    request.user.puede_aprobar = True
                else:
                    request.user.puede_aprobar = False

                request.user.save()

                # Marcar código como usado
                codigo_obj.usar_codigo(request.user)

                messages.success(
                    request,
                    f'¡Bienvenido a OIKOS! Te has unido a {codigo_obj.iglesia.nombre} como {codigo_obj.get_rol_display()}.'
                )
                return redirect('dashboard')

            # CASO 2: Código para caja chica (NUEVO)
            else:
                # Verificar que el usuario tenga iglesia asignada
                if not request.user.iglesia:
                    request.user.iglesia = codigo_obj.iglesia
                    # No asignamos rol de iglesia, el usuario solo tendrá acceso a la caja
                    # El rol por defecto es 'COLABORADOR' pero no tendrá permisos reales de iglesia
                    request.user.save()

                # Verificar si ya está asignado a esta caja
                if UsuarioCajaChica.objects.filter(
                    usuario=request.user,
                    caja_chica=codigo_obj.caja_chica
                ).exists():
                    messages.warning(request, 'Ya estás asignado a esta caja')
                    return redirect('dashboard')

                # Crear asignación a la caja
                UsuarioCajaChica.objects.create(
                    usuario=request.user,
                    caja_chica=codigo_obj.caja_chica,
                    rol_caja=codigo_obj.rol,  # TESORERO_CAJA o COLABORADOR_CAJA
                    puede_aprobar=(codigo_obj.rol == 'TESORERO_CAJA'),
                    asignado_por=codigo_obj.creado_por
                )

                # Marcar código como usado
                codigo_obj.usar_codigo(request.user)

                messages.success(
                    request,
                    f'Te has unido a la caja "{codigo_obj.caja_chica.nombre}" '
                    f'como {codigo_obj.get_rol_display()}.'
                )
                return redirect('dashboard')
    else:
        form = ValidarCodigoInvitacionForm()

    return render(request, 'core/registro_con_codigo.html', {
        'form': form,
        'user': request.user
    })


class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'core/dashboard.html'

    def dispatch(self, request, *args, **kwargs):
        # Si el usuario no tiene iglesia, redirigir a selección de tipo de registro
        if request.user.is_authenticated:
            if not request.user.is_staff and not request.user.is_superuser:
                if not request.user.iglesia:
                    return redirect('seleccionar_tipo_registro')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        iglesia = self.request.user.iglesia
        # Permitir seleccionar mes desde GET, por defecto el mes anterior al actual
        mes_anterior = (timezone.now() - relativedelta(months=1)).strftime('%Y-%m')
        mes_seleccionado = self.request.GET.get('mes', mes_anterior)

        # El saldo actual es el TOTAL acumulado de todos los movimientos hasta hoy (excluye anulados)
        from django.db.models import Sum
        total_ingresos_historico = Movimiento.objects.filter(
            iglesia=iglesia,
            tipo='INGRESO',
            anulado=False
        ).aggregate(total=Sum('monto'))['total'] or Decimal('0.00')

        total_egresos_historico = Movimiento.objects.filter(
            iglesia=iglesia,
            tipo='EGRESO',
            anulado=False
        ).aggregate(total=Sum('monto'))['total'] or Decimal('0.00')

        saldo_final = total_ingresos_historico - total_egresos_historico

        # Totales del mes SELECCIONADO (puede ser diferente al actual, excluye anulados)
        año, mes = mes_seleccionado.split('-')
        movimientos_mes = Movimiento.objects.filter(
            iglesia=iglesia,
            fecha__year=int(año),
            fecha__month=int(mes),
            anulado=False
        )

        total_ingresos_mes = movimientos_mes.filter(tipo='INGRESO').aggregate(
            total=Sum('monto')
        )['total'] or Decimal('0.00')

        total_egresos_mes = movimientos_mes.filter(tipo='EGRESO').aggregate(
            total=Sum('monto')
        )['total'] or Decimal('0.00')

        # Últimos movimientos
        ultimos_movimientos = Movimiento.objects.filter(
            iglesia=iglesia
        ).order_by('-fecha', '-fecha_creacion')[:5]

        # Alertas
        alertas = []

        if saldo_final < Decimal('2000000'):
            alertas.append({
                'tipo': 'warning',
                'mensaje': f'Saldo actual bajo: {formato_pesos(saldo_final)}'
            })

        if total_egresos_mes > total_ingresos_mes * Decimal('1.2'):
            alertas.append({
                'tipo': 'danger',
                'mensaje': 'Los egresos superan el 120% de los ingresos del mes'
            })

        # Determinar clase de color para saldo
        if saldo_final < 0:
            saldo_clase = 'saldo-negativo'
        elif saldo_final < Decimal('2000000'):
            saldo_clase = 'saldo-alerta'
        else:
            saldo_clase = 'saldo-positivo'

        # Formatear nombre del mes seleccionado en español
        fecha_sel = datetime.strptime(mes_seleccionado, '%Y-%m')
        mes_nombre = formato_mes(fecha_sel, corto=False)

        context.update({
            'iglesia': iglesia,
            'mes_seleccionado': mes_seleccionado,
            'mes_nombre': mes_nombre,
            'saldo_actual': saldo_final,
            'saldo_actual_format': formato_pesos(saldo_final),
            'saldo_clase': saldo_clase,
            'total_ingresos_mes': formato_pesos(total_ingresos_mes),
            'total_egresos_mes': formato_pesos(total_egresos_mes),
            'ultimos_movimientos': ultimos_movimientos,
            'alertas': alertas,
        })

        # Agregar información de cajas chicas
        from core.models import CajaChica

        # Si es usuario solo de caja (no tiene rol de iglesia), mostrar solo sus cajas
        if self.request.user.es_usuario_solo_caja:
            context['mostrar_movimientos'] = False
            context['cajas_usuario'] = self.request.user.cajas_asignadas.select_related('caja_chica').all()
            context['es_usuario_caja'] = True
        else:
            context['mostrar_movimientos'] = True
            context['es_usuario_caja'] = False

        # Si es ADMIN, mostrar todas las cajas de la iglesia
        if self.request.user.rol == 'ADMIN':
            cajas_chicas = CajaChica.objects.filter(
                iglesia=iglesia,
                activa=True
            )[:5]  # Mostrar solo las primeras 5

            # Agregar saldo actual de cada caja
            for caja in cajas_chicas:
                caja.saldo_actual = caja.calcular_saldo_actual()

            context['cajas_chicas'] = cajas_chicas
            context['puede_gestionar_cajas'] = True

        return context


class MovimientoCreateView(AccesoMovimientosRequiredMixin, LoginRequiredMixin, CreateView):
    model = Movimiento
    form_class = MovimientoForm
    template_name = 'core/movimiento_form.html'
    success_url = reverse_lazy('movimiento_list')

    def dispatch(self, request, *args, **kwargs):
        # El AccesoMovimientosRequiredMixin ya verifica iglesia y acceso a movimientos
        # Verificar adicionalmente que el usuario tenga permiso para crear
        if request.user.is_authenticated and not request.user.is_staff and not request.user.is_superuser:
            if not request.user.puede_crear_movimientos:
                messages.error(request, 'No tiene permisos para crear movimientos')
                return redirect('movimiento_list')
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['iglesia'] = self.request.user.iglesia
        return kwargs

    def form_valid(self, form):
        form.instance.iglesia = self.request.user.iglesia
        form.instance.creado_por = self.request.user
        response = super().form_valid(form)
        messages.success(
            self.request,
            f'Movimiento registrado exitosamente: {self.object.get_tipo_display()} de ${self.object.monto:,.2f}'
        )

        # Si el usuario clickeó "Guardar y agregar otro", redirigir con parámetro
        action = self.request.POST.get('action', 'save')
        if action == 'save_and_new':
            from django.urls import reverse
            return redirect(reverse('movimiento_list') + '?nuevo=1')

        return response


class MovimientoUpdateView(AccesoMovimientosRequiredMixin, LoginRequiredMixin, UpdateView):
    model = Movimiento
    form_class = MovimientoForm
    template_name = 'core/movimiento_list.html'
    success_url = reverse_lazy('movimiento_list')

    def dispatch(self, request, *args, **kwargs):
        # El AccesoMovimientosRequiredMixin ya verifica acceso básico
        # Verificar que el usuario sea ADMIN o TESORERO
        if not request.user.rol in ['ADMIN', 'TESORERO']:
            messages.error(request, 'No tiene permisos para editar movimientos')
            return redirect('movimiento_list')

        # Verificar que el movimiento pertenezca a la iglesia del usuario
        movimiento = self.get_object()
        if movimiento.iglesia != request.user.iglesia:
            messages.error(request, 'No puede editar movimientos de otra iglesia')
            return redirect('movimiento_list')

        # Verificar que el movimiento no esté anulado
        if movimiento.anulado:
            messages.error(request, 'No puede editar un movimiento anulado')
            return redirect('movimiento_list')

        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        return Movimiento.objects.filter(
            iglesia=self.request.user.iglesia,
            anulado=False
        )

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['iglesia'] = self.request.user.iglesia
        return kwargs

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(
            self.request,
            f'Movimiento actualizado exitosamente: {self.object.get_tipo_display()} de ${self.object.monto:,.2f}'
        )
        return response

    def form_invalid(self, form):
        # Mostrar errores específicos del formulario
        error_messages = []
        for field, errors in form.errors.items():
            if field == '__all__':
                for error in errors:
                    error_messages.append(error)
            else:
                field_label = form.fields[field].label if field in form.fields else field
                for error in errors:
                    error_messages.append(f'{field_label}: {error}')

        if error_messages:
            for error_msg in error_messages:
                messages.error(self.request, error_msg)
        else:
            messages.error(
                self.request,
                'Error al actualizar el movimiento. Por favor revise los campos.'
            )
        return redirect('movimiento_list')


class MovimientoListView(AccesoMovimientosRequiredMixin, LoginRequiredMixin, ListView):
    model = Movimiento
    template_name = 'core/movimiento_list.html'
    context_object_name = 'movimientos'
    paginate_by = 20

    def get_queryset(self):
        queryset = Movimiento.objects.filter(
            iglesia=self.request.user.iglesia
        ).order_by('-fecha', '-fecha_creacion')

        # Aplicar filtros
        form = FiltroMovimientosForm(
            self.request.GET,
            iglesia=self.request.user.iglesia
        )

        if form.is_valid():
            if form.cleaned_data.get('mes'):
                mes_str = form.cleaned_data['mes']
                try:
                    año, mes = mes_str.split('-')
                    queryset = queryset.filter(
                        fecha__year=int(año),
                        fecha__month=int(mes)
                    )
                except (ValueError, AttributeError):
                    pass

            if form.cleaned_data.get('tipo'):
                queryset = queryset.filter(tipo=form.cleaned_data['tipo'])

            if form.cleaned_data.get('categoria'):
                cat = form.cleaned_data['categoria']
                if cat.startswith('ingreso_'):
                    cat_id = int(cat.split('_')[1])
                    queryset = queryset.filter(categoria_ingreso_id=cat_id)
                elif cat.startswith('egreso_'):
                    cat_id = int(cat.split('_')[1])
                    queryset = queryset.filter(categoria_egreso_id=cat_id)

            if form.cleaned_data.get('buscar'):
                queryset = queryset.filter(
                    concepto__icontains=form.cleaned_data['buscar']
                )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = FiltroMovimientosForm(
            self.request.GET,
            iglesia=self.request.user.iglesia
        )
        return context


@login_required
def reporte_mensual_view(request):
    """
    Vista para mostrar la página de reportes mensuales
    """
    # Si el usuario no tiene iglesia, redirigir a registro de iglesia
    if request.user.is_authenticated:
        if not request.user.is_staff and not request.user.is_superuser:
            if not request.user.iglesia:
                return redirect('seleccionar_tipo_registro')

    from dateutil.relativedelta import relativedelta

    fecha_actual = datetime.now()
    mes_actual = fecha_actual.strftime('%Y-%m')
    mes_anterior = (fecha_actual - relativedelta(months=1)).strftime('%Y-%m')
    mes_anterior_2 = (fecha_actual - relativedelta(months=2)).strftime('%Y-%m')

    context = {
        'mes_actual': mes_actual,
        'mes_anterior': mes_anterior,
        'mes_anterior_2': mes_anterior_2,
    }

    return render(request, 'core/reporte_mensual.html', context)


@login_required
def generar_reporte_pdf_view(request):
    """
    Vista para generar y descargar reporte PDF mensual
    """
    # Si el usuario no tiene iglesia, redirigir a registro de iglesia
    if request.user.is_authenticated:
        if not request.user.is_staff and not request.user.is_superuser:
            if not request.user.iglesia:
                return redirect('seleccionar_tipo_registro')

            # Verificar que el usuario tenga permiso para generar reportes
            if not request.user.puede_generar_reportes:
                messages.error(request, 'No tiene permisos para generar reportes')
                return redirect('dashboard')

    iglesia = request.user.iglesia
    año_mes = request.GET.get('mes', timezone.now().strftime('%Y-%m'))

    # Generar PDF
    pdf_buffer = generar_reporte_pdf(iglesia, año_mes)

    # Retornar como descarga
    response = HttpResponse(pdf_buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_{iglesia.nombre}_{año_mes}.pdf"'

    return response


@login_required
def generar_reporte_movimientos_completo_view(request):
    """
    Vista para generar y descargar reporte PDF completo de todos los movimientos con saldo acumulado
    """
    # Verificar que el usuario tenga permiso para generar reportes
    if not request.user.is_staff and not request.user.is_superuser:
        if not request.user.puede_generar_reportes:
            messages.error(request, 'No tiene permisos para generar reportes')
            return redirect('dashboard')

    from core.utils import generar_reporte_movimientos_completo_pdf
    from datetime import datetime

    iglesia = request.user.iglesia

    # Obtener parámetros de fecha (opcionales)
    fecha_desde_str = request.GET.get('fecha_desde')
    fecha_hasta_str = request.GET.get('fecha_hasta')

    fecha_desde = None
    fecha_hasta = None

    if fecha_desde_str:
        try:
            fecha_desde = datetime.strptime(fecha_desde_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    if fecha_hasta_str:
        try:
            fecha_hasta = datetime.strptime(fecha_hasta_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    # Generar PDF
    pdf_buffer = generar_reporte_movimientos_completo_pdf(iglesia, fecha_desde, fecha_hasta)

    # Retornar como descarga
    filename = f"movimientos_{iglesia.nombre.replace(' ', '_')}"
    if fecha_desde and fecha_hasta:
        filename += f"_{fecha_desde.strftime('%Y%m%d')}-{fecha_hasta.strftime('%Y%m%d')}"
    filename += ".pdf"

    response = HttpResponse(pdf_buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


@login_required
def dashboard_data_api(request):
    """
    API para obtener datos de gráficos del dashboard
    """
    # Si el usuario no tiene iglesia, retornar datos vacíos
    if request.user.is_authenticated:
        if not request.user.is_staff and not request.user.is_superuser:
            if not request.user.iglesia:
                return JsonResponse({'labels': [], 'ingresos': [], 'egresos': []})

    iglesia = request.user.iglesia
    # Obtener mes seleccionado desde GET
    mes_seleccionado = request.GET.get('mes', None)
    data = get_dashboard_data(iglesia, meses=6, mes_distribucion=mes_seleccionado)

    return JsonResponse(data)


@login_required
def exportar_excel_view(request):
    """
    Exportar movimientos a Excel
    """
    # Si el usuario no tiene iglesia, redirigir a registro de iglesia
    if request.user.is_authenticated:
        if not request.user.is_staff and not request.user.is_superuser:
            if not request.user.iglesia:
                return redirect('seleccionar_tipo_registro')

            # Verificar que el usuario tenga permiso para generar reportes
            if not request.user.puede_generar_reportes:
                messages.error(request, 'No tiene permisos para generar reportes')
                return redirect('dashboard')

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO

    iglesia = request.user.iglesia

    # Obtener filtros (excluir movimientos anulados)
    queryset = Movimiento.objects.filter(iglesia=iglesia, anulado=False).order_by('-fecha')

    # Aplicar filtros si existen
    mes = request.GET.get('mes')
    if mes:
        año, mes_num = mes.split('-')
        queryset = queryset.filter(fecha__year=int(año), fecha__month=int(mes_num))

    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Estilos para montos
    egreso_font = Font(bold=True, color="DC3545")  # Rojo
    ingreso_font = Font(bold=True, color="198754")  # Verde

    # Headers
    headers = ['Fecha', 'Tipo', 'Categoría', 'Concepto', 'Monto', 'Comprobante']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Datos
    for row, mov in enumerate(queryset, start=2):
        categoria = mov.categoria_ingreso or mov.categoria_egreso
        ws.cell(row=row, column=1, value=mov.fecha.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=2, value=mov.get_tipo_display())
        ws.cell(row=row, column=3, value=str(categoria))
        ws.cell(row=row, column=4, value=mov.concepto)

        # Monto: negativo para egresos, positivo para ingresos
        monto_cell = ws.cell(row=row, column=5)
        if mov.tipo == 'EGRESO':
            monto_cell.value = -float(mov.monto)  # Negativo
            monto_cell.font = egreso_font  # Rojo
        else:
            monto_cell.value = float(mov.monto)
            monto_cell.font = ingreso_font  # Verde

        # Formato de número con separador de miles
        monto_cell.number_format = '#,##0.00'

        ws.cell(row=row, column=6, value=mov.comprobante_nro or '')

    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Retornar como descarga
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'movimientos_{iglesia.nombre}_{timezone.now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


@login_required
def anular_movimiento_view(request, pk):
    """
    Anular un movimiento (soft delete con motivo)
    Solo usuarios con puede_aprobar=True pueden anular
    """
    from django.contrib import messages

    movimiento = get_object_or_404(Movimiento, pk=pk, iglesia=request.user.iglesia)

    # Verificar permisos
    if not request.user.puede_aprobar:
        messages.error(request, 'No tienes permisos para anular movimientos.')
        return redirect('movimiento_list')

    # Verificar que no esté ya anulado
    if movimiento.anulado:
        messages.warning(request, 'Este movimiento ya está anulado.')
        return redirect('movimiento_list')

    if request.method == 'POST':
        motivo = request.POST.get('motivo_anulacion', '').strip()

        if not motivo:
            messages.error(request, 'Debe ingresar un motivo para la anulación.')
            return redirect('movimiento_list')

        # Anular el movimiento
        movimiento.anulado = True
        movimiento.fecha_anulacion = timezone.now()
        movimiento.motivo_anulacion = motivo
        movimiento.anulado_por = request.user
        movimiento.save()

        messages.success(request, f'Movimiento {movimiento.comprobante_nro} anulado exitosamente.')
        return redirect('movimiento_list')

    return redirect('movimiento_list')


@login_required
def gestionar_usuarios_view(request):
    """
    Vista para que el ADMIN gestione usuarios y códigos de invitación de su iglesia
    """
    # Verificar que el usuario sea ADMIN
    if not request.user.puede_gestionar_usuarios:
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('dashboard')

    # Verificar que tenga iglesia
    if not request.user.iglesia:
        return redirect('seleccionar_tipo_registro')

    from core.models import CodigoInvitacion, Usuario
    from core.forms_invitacion import GenerarCodigoInvitacionForm

    iglesia = request.user.iglesia

    # Procesar generación de nuevo código
    if request.method == 'POST' and 'generar_codigo' in request.POST:
        form = GenerarCodigoInvitacionForm(request.POST)
        if form.is_valid():
            rol = form.cleaned_data['rol']
            dias_expiracion = form.cleaned_data['dias_expiracion']
            usos_maximos = form.cleaned_data['usos_maximos']

            # Crear código
            codigo = CodigoInvitacion.crear(
                iglesia=iglesia,
                rol=rol,
                creado_por=request.user,
                dias_expiracion=dias_expiracion
            )
            codigo.usos_maximos = usos_maximos
            codigo.save()

            messages.success(
                request,
                f'Código generado: <strong>{codigo.codigo}</strong> (válido por {dias_expiracion} días)',
                extra_tags='safe'
            )
            return redirect('gestionar_usuarios')
    else:
        form = GenerarCodigoInvitacionForm()

    # Procesar revocación de código
    if request.method == 'POST' and 'revocar_codigo' in request.POST:
        codigo_id = request.POST.get('codigo_id')
        try:
            codigo = CodigoInvitacion.objects.get(id=codigo_id, iglesia=iglesia)
            codigo.activo = False
            codigo.save()
            messages.success(request, f'Código {codigo.codigo} revocado exitosamente.')
        except CodigoInvitacion.DoesNotExist:
            messages.error(request, 'Código no encontrado.')
        return redirect('gestionar_usuarios')

    # Procesar cambio de rol de usuario
    if request.method == 'POST' and 'cambiar_rol' in request.POST:
        usuario_id = request.POST.get('usuario_id')
        nuevo_rol = request.POST.get('nuevo_rol')
        try:
            usuario = Usuario.objects.get(id=usuario_id, iglesia=iglesia)

            # No permitir cambiar el rol del único ADMIN
            if usuario.rol == 'ADMIN' and nuevo_rol != 'ADMIN':
                admins_count = Usuario.objects.filter(iglesia=iglesia, rol='ADMIN').count()
                if admins_count <= 1:
                    messages.error(request, 'No se puede cambiar el rol del único administrador.')
                    return redirect('gestionar_usuarios')

            usuario.rol = nuevo_rol
            usuario.save()
            messages.success(request, f'Rol de {usuario.get_full_name() or usuario.username} cambiado a {usuario.get_rol_display()}.')
        except Usuario.DoesNotExist:
            messages.error(request, 'Usuario no encontrado.')
        return redirect('gestionar_usuarios')

    # Procesar desactivar/activar usuario
    if request.method == 'POST' and 'toggle_usuario' in request.POST:
        usuario_id = request.POST.get('usuario_id')
        try:
            usuario = Usuario.objects.get(id=usuario_id, iglesia=iglesia)

            # No permitir desactivar al propio usuario
            if usuario.id == request.user.id:
                messages.error(request, 'No puedes desactivar tu propia cuenta.')
                return redirect('gestionar_usuarios')

            # No permitir desactivar al único ADMIN
            if usuario.rol == 'ADMIN' and usuario.is_active:
                admins_count = Usuario.objects.filter(iglesia=iglesia, rol='ADMIN', is_active=True).count()
                if admins_count <= 1:
                    messages.error(request, 'No se puede desactivar al único administrador activo.')
                    return redirect('gestionar_usuarios')

            usuario.is_active = not usuario.is_active
            usuario.save()

            estado = 'activado' if usuario.is_active else 'desactivado'
            messages.success(request, f'Usuario {usuario.get_full_name() or usuario.username} {estado} exitosamente.')
        except Usuario.DoesNotExist:
            messages.error(request, 'Usuario no encontrado.')
        return redirect('gestionar_usuarios')

    # Procesar eliminar usuario
    if request.method == 'POST' and 'eliminar_usuario' in request.POST:
        usuario_id = request.POST.get('usuario_id')
        try:
            usuario = Usuario.objects.get(id=usuario_id, iglesia=iglesia)

            # No permitir eliminar al propio usuario
            if usuario.id == request.user.id:
                messages.error(request, 'No puedes eliminar tu propia cuenta.')
                return redirect('gestionar_usuarios')

            # No permitir eliminar al único ADMIN
            if usuario.rol == 'ADMIN':
                admins_count = Usuario.objects.filter(iglesia=iglesia, rol='ADMIN').count()
                if admins_count <= 1:
                    messages.error(request, 'No se puede eliminar al único administrador.')
                    return redirect('gestionar_usuarios')

            # Verificar si el usuario ha creado movimientos
            movimientos_count = usuario.movimientos_creados.count()
            if movimientos_count > 0:
                messages.error(
                    request,
                    f'No se puede eliminar al usuario porque ha creado {movimientos_count} movimiento(s). '
                    f'Por razones de auditoría, los usuarios con movimientos registrados deben ser desactivados en lugar de eliminados.'
                )
                return redirect('gestionar_usuarios')

            nombre_usuario = usuario.get_full_name() or usuario.username
            usuario.delete()
            messages.success(request, f'Usuario {nombre_usuario} eliminado exitosamente.')
        except Usuario.DoesNotExist:
            messages.error(request, 'Usuario no encontrado.')
        return redirect('gestionar_usuarios')

    # Obtener datos para el template
    usuarios = Usuario.objects.filter(iglesia=iglesia).order_by('-date_joined')

    codigos_activos = CodigoInvitacion.objects.filter(
        iglesia=iglesia,
        activo=True
    ).filter(
        fecha_expiracion__gt=timezone.now()
    ).order_by('-fecha_creacion')

    codigos_usados = CodigoInvitacion.objects.filter(
        iglesia=iglesia,
        usado_por__isnull=False
    ).order_by('-fecha_uso')

    codigos_expirados = CodigoInvitacion.objects.filter(
        iglesia=iglesia
    ).exclude(
        id__in=codigos_activos.values_list('id', flat=True)
    ).exclude(
        id__in=codigos_usados.values_list('id', flat=True)
    ).order_by('-fecha_creacion')

    context = {
        'form': form,
        'usuarios': usuarios,
        'codigos_activos': codigos_activos,
        'codigos_usados': codigos_usados,
        'codigos_expirados': codigos_expirados,
        'total_usuarios': usuarios.count(),
    }

    return render(request, 'core/gestionar_usuarios.html', context)

@login_required
def ayuda_view(request):
    """
    Vista de ayuda con explicación de métricas del dashboard
    """
    return render(request, 'core/ayuda.html')


@login_required
def contadora_billetes_view(request):
    """
    Vista de la contadora de billetes para facilitar el conteo de efectivo
    """
    return render(request, 'core/contadora_billetes.html')

def politica_cookies_view(request):
    """
    Vista para mostrar la política de cookies
    """
    return render(request, 'core/politica_cookies.html')


def terminos_condiciones_view(request):
    """
    Vista para mostrar los términos y condiciones
    """
    return render(request, 'core/terminos_condiciones.html')


@login_required
def aceptar_terminos_view(request):
    """
    Vista para procesar la aceptación de términos y condiciones
    """
    if request.method == 'POST':
        from django.utils import timezone

        # Registrar aceptación
        request.user.terminos_aceptados = True
        request.user.fecha_aceptacion_terminos = timezone.now()
        request.user.save()

        messages.success(request, 'Has aceptado los términos y condiciones exitosamente.')

        # Redirigir a donde venía o al dashboard
        next_url = request.POST.get('next', 'dashboard')
        return redirect(next_url)

    # Si no es POST, redirigir al dashboard
    return redirect('dashboard')


# ============================================================================
# VISTAS DE CATEGORÍAS DE INGRESO
# ============================================================================

class CategoriaIngresoListView(AccesoMovimientosRequiredMixin, LoginRequiredMixin, ListView):
    model = CategoriaIngreso
    template_name = 'core/categoria_ingreso_list.html'
    context_object_name = 'categorias'
    paginate_by = 20

    def get_queryset(self):
        return CategoriaIngreso.objects.filter(
            iglesia=self.request.user.iglesia
        ).order_by('codigo')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['activas'] = self.get_queryset().filter(activa=True).count()
        context['inactivas'] = self.get_queryset().filter(activa=False).count()
        return context


class CategoriaIngresoCreateView(AccesoMovimientosRequiredMixin, LoginRequiredMixin, CreateView):
    model = CategoriaIngreso
    form_class = CategoriaIngresoForm
    template_name = 'core/categoria_ingreso_list.html'
    success_url = reverse_lazy('categoria_ingreso_list')

    def dispatch(self, request, *args, **kwargs):
        # El AccesoMovimientosRequiredMixin ya verifica acceso básico
        # Solo ADMIN puede crear categorías
        if request.user.rol != 'ADMIN':
            messages.error(request, 'Solo los administradores pueden crear categorías.')
            return redirect('categoria_ingreso_list')
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['iglesia'] = self.request.user.iglesia
        return kwargs

    def form_valid(self, form):
        form.instance.iglesia = self.request.user.iglesia
        response = super().form_valid(form)
        messages.success(
            self.request,
            f'Categoría "{self.object.nombre}" creada exitosamente'
        )
        return response

    def form_invalid(self, form):
        messages.error(
            self.request,
            'Error al crear la categoría. Por favor revise los campos.'
        )
        return redirect('categoria_ingreso_list')


class CategoriaIngresoUpdateView(AccesoMovimientosRequiredMixin, LoginRequiredMixin, UpdateView):
    model = CategoriaIngreso
    form_class = CategoriaIngresoForm
    template_name = 'core/categoria_ingreso_list.html'
    success_url = reverse_lazy('categoria_ingreso_list')

    def dispatch(self, request, *args, **kwargs):
        # El AccesoMovimientosRequiredMixin ya verifica acceso básico
        # Solo ADMIN puede modificar categorías
        if request.user.rol != 'ADMIN':
            messages.error(request, 'Solo los administradores pueden modificar categorías.')
            return redirect('categoria_ingreso_list')
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        return CategoriaIngreso.objects.filter(iglesia=self.request.user.iglesia)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['iglesia'] = self.request.user.iglesia
        return kwargs

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(
            self.request,
            f'Categoría "{self.object.nombre}" actualizada exitosamente'
        )
        return response

    def form_invalid(self, form):
        messages.error(
            self.request,
            'Error al actualizar la categoría. Por favor revise los campos.'
        )
        return redirect('categoria_ingreso_list')


@login_required
def toggle_categoria_ingreso(request, pk):
    """Activa o desactiva una categoría de ingreso"""
    # Solo ADMIN puede activar/desactivar categorías
    if request.user.rol != 'ADMIN':
        messages.error(request, 'Solo los administradores pueden activar/desactivar categorías.')
        return redirect('categoria_ingreso_list')

    categoria = get_object_or_404(
        CategoriaIngreso,
        pk=pk,
        iglesia=request.user.iglesia
    )

    categoria.activa = not categoria.activa
    categoria.save()

    estado = "activada" if categoria.activa else "desactivada"
    messages.success(
        request,
        f'Categoría "{categoria.nombre}" {estado} exitosamente'
    )

    return redirect('categoria_ingreso_list')


# ============================================================================
# VISTAS DE CATEGORÍAS DE EGRESO
# ============================================================================

class CategoriaEgresoListView(AccesoMovimientosRequiredMixin, LoginRequiredMixin, ListView):
    model = CategoriaEgreso
    template_name = 'core/categoria_egreso_list.html'
    context_object_name = 'categorias'
    paginate_by = 20

    def get_queryset(self):
        return CategoriaEgreso.objects.filter(
            iglesia=self.request.user.iglesia
        ).order_by('codigo')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['activas'] = self.get_queryset().filter(activa=True).count()
        context['inactivas'] = self.get_queryset().filter(activa=False).count()
        return context


class CategoriaEgresoCreateView(AccesoMovimientosRequiredMixin, LoginRequiredMixin, CreateView):
    model = CategoriaEgreso
    form_class = CategoriaEgresoForm
    template_name = 'core/categoria_egreso_list.html'
    success_url = reverse_lazy('categoria_egreso_list')

    def dispatch(self, request, *args, **kwargs):
        # Solo ADMIN puede crear categorías
        if request.user.rol != 'ADMIN':
            messages.error(request, 'Solo los administradores pueden crear categorías.')
            return redirect('categoria_egreso_list')
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['iglesia'] = self.request.user.iglesia
        return kwargs

    def form_valid(self, form):
        form.instance.iglesia = self.request.user.iglesia
        response = super().form_valid(form)
        messages.success(
            self.request,
            f'Categoría "{self.object.nombre}" creada exitosamente'
        )
        return response

    def form_invalid(self, form):
        messages.error(
            self.request,
            'Error al crear la categoría. Por favor revise los campos.'
        )
        return redirect('categoria_egreso_list')


class CategoriaEgresoUpdateView(AccesoMovimientosRequiredMixin, LoginRequiredMixin, UpdateView):
    model = CategoriaEgreso
    form_class = CategoriaEgresoForm
    template_name = 'core/categoria_egreso_list.html'
    success_url = reverse_lazy('categoria_egreso_list')

    def dispatch(self, request, *args, **kwargs):
        # Solo ADMIN puede modificar categorías
        if request.user.rol != 'ADMIN':
            messages.error(request, 'Solo los administradores pueden modificar categorías.')
            return redirect('categoria_egreso_list')
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        return CategoriaEgreso.objects.filter(iglesia=self.request.user.iglesia)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['iglesia'] = self.request.user.iglesia
        return kwargs

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(
            self.request,
            f'Categoría "{self.object.nombre}" actualizada exitosamente'
        )
        return response

    def form_invalid(self, form):
        messages.error(
            self.request,
            'Error al actualizar la categoría. Por favor revise los campos.'
        )
        return redirect('categoria_egreso_list')


@login_required
def toggle_categoria_egreso(request, pk):
    """Activa o desactiva una categoría de egreso"""
    # Solo ADMIN puede activar/desactivar categorías
    if request.user.rol != 'ADMIN':
        messages.error(request, 'Solo los administradores pueden activar/desactivar categorías.')
        return redirect('categoria_egreso_list')

    categoria = get_object_or_404(
        CategoriaEgreso,
        pk=pk,
        iglesia=request.user.iglesia
    )

    categoria.activa = not categoria.activa
    categoria.save()

    estado = "activada" if categoria.activa else "desactivada"
    messages.success(
        request,
        f'Categoría "{categoria.nombre}" {estado} exitosamente'
    )

    return redirect('categoria_egreso_list')


@login_required
def perfil_usuario_view(request):
    """
    Vista para mostrar el perfil del usuario con toda su información
    """
    usuario = request.user

    # Obtener cajas asignadas si las tiene
    cajas_asignadas = usuario.cajas_asignadas.select_related('caja_chica').all() if hasattr(usuario, 'cajas_asignadas') else []

    # Obtener estadísticas de movimientos creados (si tiene acceso)
    movimientos_creados = 0
    total_ingresos_creados = Decimal('0.00')
    total_egresos_creados = Decimal('0.00')

    if usuario.tiene_acceso_movimientos:
        movimientos = Movimiento.objects.filter(
            creado_por=usuario,
            anulado=False,
            iglesia=usuario.iglesia
        )
        movimientos_creados = movimientos.count()
        total_ingresos_creados = movimientos.filter(tipo='INGRESO').aggregate(
            total=Sum('monto')
        )['total'] or Decimal('0.00')
        total_egresos_creados = movimientos.filter(tipo='EGRESO').aggregate(
            total=Sum('monto')
        )['total'] or Decimal('0.00')

    contexto = {
        'usuario': usuario,
        'cajas_asignadas': cajas_asignadas,
        'movimientos_creados': movimientos_creados,
        'total_ingresos_creados': total_ingresos_creados,
        'total_egresos_creados': total_egresos_creados,
    }

    return render(request, 'core/perfil_usuario.html', contexto)
